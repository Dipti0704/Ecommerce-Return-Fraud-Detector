"""
reports/excel_writer.py
=======================
Produces the three-sheet Excel fraud report.

Column list + display names come from config/settings.py (EXCEL_COLUMNS).
Styling colours come from config/settings.py (EXCEL_STYLE).
No column names or colours are hardcoded here.
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import EXCEL_COLUMNS, EXCEL_STYLE as ES, CURRENCY_SYMBOL
from utils.schema import FeatSchema as FS
from utils.logger import get_logger

log = get_logger("reports.excel_writer")

# ── Style helpers ─────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(hex_color: str, bold: bool = False, size: int = 10) -> Font:
    return Font(color=hex_color, bold=bold, size=size)

_THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
    right =Side(style="thin", color="EEEEEE"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# Columns that get currency format
_CURRENCY_COLS = {"total_spend", "total_refund_amt", "avg_refund_amt", "max_order_value"}
# Columns that get percentage format
_PCT_COLS      = {"return_rate"}


class ExcelReportWriter:
    """
    Builds and saves the fraud detection Excel workbook.

    Usage
    -----
        writer = ExcelReportWriter()
        writer.write(scored_df, orders_df, returns_df, output_path)
    """

    def write(
        self,
        scored_df : pd.DataFrame,
        orders_df : pd.DataFrame,
        returns_df: pd.DataFrame,
        output_path: str | Path,
    ):
        log.info("Building Excel report → %s", output_path)
        wb = Workbook()

        # Derive column lists from config — no hardcoding
        col_keys    = [c[0] for c in EXCEL_COLUMNS]   # internal names
        col_headers = [c[1] for c in EXCEL_COLUMNS]   # display names
        col_widths  = [c[2] for c in EXCEL_COLUMNS]   # widths

        # Only keep columns that actually exist in the dataframe
        present = [c for c in col_keys if c in scored_df.columns]
        headers = [col_headers[col_keys.index(c)] for c in present]
        widths  = [col_widths [col_keys.index(c)] for c in present]

        sorted_df = scored_df.sort_values(FS.FRAUD_SCORE, ascending=False).reset_index(drop=True)

        self._sheet_all_users(wb.active, sorted_df, present, headers, widths)
        self._sheet_high_risk(wb.create_sheet("🚨 High Risk Users"), sorted_df, present, headers, widths)
        self._sheet_summary  (wb.create_sheet("Summary & KPIs"),     scored_df, orders_df, returns_df)

        wb.save(output_path)
        log.info("Excel saved: %s", Path(output_path).name)

    # ── Sheet 1: All users sorted by fraud score ──────────────────────────

    def _sheet_all_users(self, ws, df, cols, headers, widths):
        ws.title = "Fraud Risk Report"
        self._write_header(ws, headers, ES["header_fill"], ES["header_font"])
        self._set_col_widths(ws, widths)

        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            risk      = row[FS.RISK_LEVEL]
            row_fill  = _fill(self._risk_row_fill(risk))
            ws.row_dimensions[row_idx].height = 20

            for col_idx, col in enumerate(cols, 1):
                val  = self._cell_value(row, col)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill   = row_fill
                cell.border = _THIN_BORDER
                self._apply_cell_style(cell, col, risk)

        ws.freeze_panes = "A2"

    # ── Sheet 2: High-risk users only ─────────────────────────────────────

    def _sheet_high_risk(self, ws, df, cols, headers, widths):
        high_df = df[df[FS.RISK_LEVEL] == "High"].copy()
        self._write_header(ws, headers, ES["hr_header_fill"], "FFFFFF")
        self._set_col_widths(ws, widths)

        for row_idx, (_, row) in enumerate(high_df.iterrows(), 2):
            alt_fill = _fill("FFF0F0" if row_idx % 2 == 0 else "FFFFFF")
            for col_idx, col in enumerate(cols, 1):
                val  = self._cell_value(row, col)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill   = alt_fill
                cell.border = _THIN_BORDER
                cell.alignment = _LEFT if col == FS.FLAGGED_REASONS else _CENTER
                if col in _PCT_COLS:     cell.number_format = "0.0%"
                if col in _CURRENCY_COLS:cell.number_format = "#,##0.00"
                if col == FS.FRAUD_SCORE:
                    cell.font = _font(ES["high_score_font"], bold=True)

        ws.freeze_panes = "A2"

    # ── Sheet 3: Summary KPIs ─────────────────────────────────────────────

    def _sheet_summary(self, ws, scored_df, orders_df, returns_df):
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 22

        risk_counts = scored_df[FS.RISK_LEVEL].value_counts()
        high_df     = scored_df[scored_df[FS.RISK_LEVEL] == "High"]
        n_users     = len(scored_df)
        n_orders    = len(orders_df)
        n_returns   = len(returns_df)
        ret_rate    = n_returns / n_orders if n_orders else 0

        cur = CURRENCY_SYMBOL

        kpis = [
            ("OVERVIEW",                       "",  ""),
            ("Total users analysed",           n_users,  ""),
            ("Total orders processed",         n_orders, ""),
            ("Total returns",                  n_returns,""),
            ("Overall return rate",            f"{ret_rate:.1%}", ""),
            ("", "", ""),
            ("RISK DISTRIBUTION",              "", ""),
            ("High risk users",                int(risk_counts.get("High",   0)),
             f"{risk_counts.get('High',0)/n_users:.1%} of users"),
            ("Medium risk users",              int(risk_counts.get("Medium", 0)),
             f"{risk_counts.get('Medium',0)/n_users:.1%} of users"),
            ("Low risk users",                 int(risk_counts.get("Low",    0)),
             f"{risk_counts.get('Low',0)/n_users:.1%} of users"),
            ("", "", ""),
            ("FINANCIAL EXPOSURE",             "", ""),
            ("Total spend — high risk",
             f"{cur}{high_df['total_spend'].sum():,.0f}",       ""),
            ("Total refunds — high risk",
             f"{cur}{high_df['total_refund_amt'].sum():,.0f}",  ""),
            ("Avg fraud score — high risk",
             f"{high_df[FS.FRAUD_SCORE].mean():.1f} / 100",    ""),
            ("Avg return rate — high risk",
             f"{high_df['return_rate'].mean():.1%}",            ""),
            ("", "", ""),
            ("MODEL PERFORMANCE",              "", ""),
            ("Algorithm",    "Random Forest + Isolation Forest Ensemble", ""),
            ("Ensemble weights", "RF 65% · IsolationForest 35%", ""),
            ("Features used",    len([c for c in scored_df.columns
                                      if c not in [FS.FRAUD_SCORE, FS.RISK_LEVEL,
                                                   FS.FLAGGED_REASONS, FS.RF_PROBABILITY,
                                                   FS.ANOMALY_SCORE]]), ""),
        ]

        SECTION_FILL = _fill(ES["section_fill"])
        SECTION_FONT = _font(ES["section_font"], bold=True)

        for row_idx, (label, value, note) in enumerate(kpis, 2):
            ws.row_dimensions[row_idx].height = 22
            la = ws.cell(row=row_idx, column=1, value=label)
            va = ws.cell(row=row_idx, column=2, value=value)
            na = ws.cell(row=row_idx, column=3, value=note)
            if label in ("OVERVIEW", "RISK DISTRIBUTION", "FINANCIAL EXPOSURE", "MODEL PERFORMANCE"):
                for cell in (la, va, na):
                    cell.fill = SECTION_FILL
                la.font = SECTION_FONT
            elif label:
                la.font = _font("333333")
                va.font = _font("1A237E", bold=True)
                na.font = _font("888888", size=9)

    # ── Shared utilities ──────────────────────────────────────────────────

    @staticmethod
    def _write_header(ws, headers: list, fill_hex: str, font_hex: str):
        ws.row_dimensions[1].height = 28
        for col_idx, header in enumerate(headers, 1):
            cell            = ws.cell(row=1, column=col_idx, value=header)
            cell.fill       = _fill(fill_hex)
            cell.font       = _font(font_hex, bold=True)
            cell.border     = _THIN_BORDER
            cell.alignment  = _CENTER

    @staticmethod
    def _set_col_widths(ws, widths: list):
        for idx, w in enumerate(widths, 1):
            if w:
                ws.column_dimensions[get_column_letter(idx)].width = w

    @staticmethod
    def _cell_value(row: pd.Series, col: str):
        val = row[col]
        if col in _PCT_COLS:
            return float(val)
        return val

    def _risk_row_fill(self, risk: str) -> str:
        mapping = {
            "High"  : ES["high_row_fill"],
            "Medium": ES["med_row_fill"],
            "Low"   : ES["low_row_fill"],
        }
        return mapping.get(risk, "FFFFFF")

    def _apply_cell_style(self, cell, col: str, risk: str):
        if col == FS.FRAUD_SCORE:
            color_map = {
                "High"  : ES["high_score_font"],
                "Medium": ES["med_score_font"],
                "Low"   : ES["low_score_font"],
            }
            cell.font      = _font(color_map.get(risk, "000000"), bold=True)
            cell.alignment = _CENTER

        elif col == FS.RISK_LEVEL:
            fill_map = {
                "High"  : (ES["high_badge_fill"], ES["high_badge_font"]),
                "Medium": (ES["med_badge_fill"],  ES["med_badge_font"]),
                "Low"   : (ES["low_badge_fill"],  ES["low_badge_font"]),
            }
            bg, fg    = fill_map.get(risk, ("FFFFFF", "000000"))
            cell.fill       = _fill(bg)
            cell.font       = _font(fg, bold=True, size=9)
            cell.alignment  = _CENTER

        elif col in _PCT_COLS:
            cell.number_format = "0.0%"
            cell.alignment     = _CENTER

        elif col in _CURRENCY_COLS:
            cell.number_format = "#,##0.00"
            cell.alignment     = _CENTER

        elif col == FS.FLAGGED_REASONS:
            cell.alignment = _LEFT
            cell.font      = _font(ES["reasons_font"], size=9)

        else:
            cell.alignment = _CENTER
